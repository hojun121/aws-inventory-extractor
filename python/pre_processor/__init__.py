import sys
import os
import re
import time
import pandas as pd
import openpyxl
from datetime import datetime
from sqlalchemy import create_engine
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from tqdm import tqdm

from modules.vpc import load_and_transform_vpc_data
from modules.tgw import load_and_transform_tgw_data
from modules.vep import load_and_transform_vep_data
from modules.pc import load_and_transform_pc_data
from modules.subnet import load_and_transform_subnet_data
from modules.security_groups import load_and_transform_security_groups_data
from modules.autoscaling import load_and_transform_autoscaling_data
from modules.cloudfront import load_and_transform_cloudfront_data
from modules.docdb import load_and_transform_docdb_data
from modules.ec import load_and_transform_elasticache_data
from modules.ec2 import load_and_transform_ec2_data
from modules.elb import load_and_transform_elb_data
from modules.iamrole import load_and_transform_iam_role_data
from modules.iamuser import load_and_transform_iam_user_data
from modules.iamgroup import load_and_transform_iam_group_data
from modules.nacl import load_and_transform_nacl_data
from modules.s3 import load_and_transform_s3_data
from modules.tg import load_and_transform_target_group_data
from modules.rds import load_and_transform_rds_data

def main():
    # Input password from command
    # password=$(steampipe service start --show-password | grep 'Password:' | awk '{print $2}')
    if len(sys.argv) < 2 or sys.argv[1] == "None":
        print("No input or 'None' provided. Exiting the program.")
        sys.exit(1)

    password = sys.argv[1]
    DB_URI = f'postgresql://steampipe:{password}@127.0.0.1:9193/steampipe'
    print(DB_URI)

    try:
        engine = create_engine(DB_URI)
        connection = engine.connect()
        print("Connection successful.")
    except Exception as e:
        print("Please check your SSO or IAM permissions.")
        sys.exit(1)

    today = datetime.today().strftime('%y%m%d')
    schemas = get_schemas()
    for schema in schemas:
        output_excel_path = os.path.join('/app/output/pre_processed', f'{schema}_inventory_{today}.xlsx')
        os.makedirs('/app/output/pre_processed', exist_ok=True)
        queries = get_queries(schema)
        process_and_save_sheets(queries, output_excel_path, engine, schema)

def get_schemas():
    schemas = []
    try:
        with open(os.path.expanduser("~/.steampipe/config/aws.spc"), "r") as file:
            content = file.read()
            matches = re.findall(r'connection\s+"([^"]+)"', content)
            if matches:
                schemas.extend(matches)
    except FileNotFoundError:
        print("aws.spc File not found")
    return schemas

def get_queries(schema):
    return {
        'alb': f'SELECT * FROM {schema}.aws_ec2_application_load_balancer',
        'autoscaling': f'SELECT * FROM {schema}.aws_ec2_autoscaling_group',
        'cloudfront': f'SELECT * FROM {schema}.aws_cloudfront_distribution',
        'cloudwatch': f'SELECT * FROM {schema}.aws_cloudwatch_metric',
        'docdbcluster': f'SELECT * FROM {schema}.aws_docdb_cluster',
        'docdbinstance': f'SELECT * FROM {schema}.aws_docdb_cluster_instance',
        'ebs': f'SELECT * FROM {schema}.aws_ebs_volume',
        'ec': f'SELECT * FROM {schema}.aws_elasticache_cluster',
        'ecrep': f'SELECT * FROM {schema}.aws_elasticache_replication_group',
        'ec2': f'SELECT * FROM {schema}.aws_ec2_instance',
        'igw': f'SELECT * FROM {schema}.aws_vpc_internet_gateway',
        'iamgroup': f'SELECT * FROM {schema}.aws_iam_group',
        'iamrole': f'SELECT * FROM {schema}.aws_iam_role',
        'iamuser': f'SELECT * FROM {schema}.aws_iam_user',
        'lbl': f'SELECT * FROM {schema}.aws_ec2_load_balancer_listener',
        'ngw': f'SELECT * FROM {schema}.aws_vpc_nat_gateway',
        'nacl': f'SELECT * FROM {schema}.aws_vpc_network_acl',
        'nlb': f'SELECT * FROM {schema}.aws_ec2_network_load_balancer',
        'pc': f'SELECT * FROM {schema}.aws_vpc_peering_connection',
        'rdscluster': f'SELECT * FROM {schema}.aws_rds_db_cluster',
        'rdsinstance': f'SELECT * FROM {schema}.aws_rds_db_instance',
        'rt': f'SELECT * FROM {schema}.aws_vpc_route_table',
        'sg': f'SELECT * FROM {schema}.aws_vpc_security_group',
        'sgrule': f'SELECT * FROM {schema}.aws_vpc_security_group_rule',
        's3': f'SELECT * FROM {schema}.aws_s3_bucket',
        'subnet': f'SELECT * FROM {schema}.aws_vpc_subnet',
        'tg': f'SELECT * FROM {schema}.aws_ec2_target_group',
        'tgw': f'SELECT * FROM {schema}.aws_ec2_transit_gateway',
        'vep': f'SELECT * FROM {schema}.aws_vpc_endpoint',
        'vpc': f'SELECT * FROM {schema}.aws_vpc',
    }

def process_and_save_sheets(queries, output_excel_path, engine, schema):
    try:
        data_dict = {query: pd.read_sql(queries[query], engine) for query in tqdm(queries, desc=f"Executing SQL Queries: {schema}")}
        with pd.ExcelWriter(output_excel_path, engine='xlsxwriter') as writer:
            transformation_tasks = [
                ('VPC', load_and_transform_vpc_data, [data_dict['vpc'], data_dict['igw'], data_dict['ngw']]),
                ('VPC Endpoint', load_and_transform_vep_data, [data_dict['vep']]),
                ('Peering Connection', load_and_transform_pc_data, [data_dict['pc']]),
                ('Transit Gateway', load_and_transform_tgw_data, [data_dict['tgw']]),
                ('Subnet', load_and_transform_subnet_data, [data_dict['subnet'], data_dict['rt'], data_dict['nacl']]),
                ('Security Groups', load_and_transform_security_groups_data, [data_dict['sg'], data_dict['sgrule']]),
                ('Network ACLs', load_and_transform_nacl_data, [data_dict['nacl']]),
                ('EC2', load_and_transform_ec2_data, [data_dict['ec2'], data_dict['ebs']]),
                ('ELB', load_and_transform_elb_data, [data_dict['alb'], data_dict['nlb'], data_dict['lbl']]),
                ('Target Group', load_and_transform_target_group_data, [data_dict['tg'], data_dict['autoscaling'], data_dict['ec2']]),
                ('Auto Scaling', load_and_transform_autoscaling_data, [data_dict['autoscaling']]),
                ('ElastiCache', load_and_transform_elasticache_data, [data_dict['ec'], data_dict['ecrep']]),
                ('CloudFront', load_and_transform_cloudfront_data, [data_dict['cloudfront']]),
                ('S3', load_and_transform_s3_data, [data_dict['s3']]),
                ('IAM Group', load_and_transform_iam_group_data, [data_dict['iamgroup']]),
                ('IAM Role', load_and_transform_iam_role_data, [data_dict['iamrole']]),
                ('IAM User', load_and_transform_iam_user_data, [data_dict['iamuser']]),
                ('RDS', load_and_transform_rds_data, [data_dict['rdscluster'], data_dict['rdsinstance'], data_dict['cloudwatch']]),
                ('DocumentDB', load_and_transform_docdb_data, [data_dict['docdbcluster'], data_dict['docdbinstance'], data_dict['cloudwatch']])
            ]

            for sheet_name, transform_function, params in tqdm(transformation_tasks, desc="Transforming and Saving Data"):
                if not params[0].empty:
                    transformed_data = transform_function(*params)
                    transformed_data.to_excel(writer, sheet_name=sheet_name, index=False)
        excel_styler(output_excel_path)
    except Exception as e:
        print(f"__init__.py > process_and_save_sheets(): {e}")

def excel_styler(output_excel_path):
    wb = openpyxl.load_workbook(output_excel_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        excel_sheet_styler(ws)
    wb.save(output_excel_path)

def excel_sheet_styler(sheet):
    # adjust_column_widths
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        column_name = column_cells[0].value

        for cell in column_cells:
            try:
                if cell.value:
                    cell_value = str(cell.value).split("\n")
                    max_cell_length = max(len(line) for line in cell_value)
                    max_length = max(max_length, max_cell_length)
            except Exception as e:
                print(f"Error calculating length for cell {cell.coordinate}: {e}")

        if column_name == 'Tag':
            adjusted_width = 50
        elif column_name == 'Listener From':
            adjusted_width = 15
        elif column_name == 'Listener To':
            adjusted_width = 30
        else:
            adjusted_width = max_length + 2

        sheet.column_dimensions[column_letter].width = adjusted_width
    # style_header
    header_fill = PatternFill(start_color="334d1d", end_color="334d1d", fill_type="solid")
    header_font = Font(color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center')
    # apply_borders
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
    # center_align_cells
    for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=1), start=1):
        alignment = Alignment(horizontal='right', vertical='center') if col[0].value == "Tag" else Alignment(vertical='center')
        for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = alignment

if __name__ == '__main__':
    main()


