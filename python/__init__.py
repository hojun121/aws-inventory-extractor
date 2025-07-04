import boto3
import pandas as pd
import os
import re
import sys
from datetime import datetime
from botocore.exceptions import ProfileNotFound, NoCredentialsError, ClientError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from modules.ec2 import list_ec2_instances

def list_all_regions(session):
    ec2 = session.client("ec2")
    response = ec2.describe_regions(AllRegions=True)
    return [region['RegionName'] for region in response['Regions'] if region.get('OptInStatus', 'opt-in-not-required') in ['opt-in-not-required', 'opted-in']]

def get_account_id(session):
    sts_client = session.client('sts')
    identity = sts_client.get_caller_identity()
    return identity['Account']

def get_aws_profiles(aws_config_path):
    config_path = os.path.expanduser(aws_config_path)
    if not os.path.exists(config_path):
        print(f"Config file not found at {config_path}")
        sys.exit(1)

    with open(config_path, 'r') as file:
        config_content = file.read()
        profiles = re.findall(r'\[profile (.*?)\]', config_content)
        return profiles

def workbook_with_format(file_name_with_dir):
    try:
        workbook = load_workbook(file_name_with_dir)
        header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.auto_filter.ref = worksheet.dimensions
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                column_letter = get_column_letter(column[0].column)
                worksheet.column_dimensions[column_letter].width = max_length
                for cell in column:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(file_name_with_dir)
    except Exception as e:
        print(f"Error formatting workbook: {e}")

def collect_all_resources():
    profile_names = get_aws_profiles("~/.aws/config")
    all_data = []

    for profile in profile_names:
        print(f"Processing profile: {profile}")
        try:
            session = boto3.Session(profile_name=profile)
            account_id = get_account_id(session)
            regions = list_all_regions(session)

            for region in regions:
                print(f"  Fetching EC2 instances in region: {region}")
                data = list_ec2_instances(session, region, profile, account_id)
                all_data.extend(data)

        except (ProfileNotFound, NoCredentialsError, ClientError) as e:
            print(f"Error accessing profile {profile}: {e}")

    df = pd.DataFrame(all_data)
    output_dir = "./"
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"aws_ec2_inventory_{timestamp}.xlsx")
    df.to_excel(output_file, index=False)

    workbook_with_format(output_file)

    print(f"EC2 inventory saved to {output_file}")

if __name__ == "__main__":
    collect_all_resources()
