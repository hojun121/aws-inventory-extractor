from flask import Flask, render_template, jsonify, request, send_file
import boto3
import re
import os
import json
import subprocess
import configparser
import zipfile
import pandas as pd
from datetime import datetime, timezone
from tempfile import NamedTemporaryFile
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl import load_workbook
from modules.vpc import list_vpcs
from modules.subnet import list_subnets
from modules.nacl import list_nacls
from modules.ec2 import list_ec2_instances
from modules.asg import list_auto_scaling_groups
from modules.elb import list_elbs
from modules.tg import list_target_groups
from modules.cloudfront import list_cloudfront_distributions
from modules.s3 import list_s3_buckets
from modules.iamrole import list_iam_roles
from modules.db import list_db_clusters
from modules.elasticache import list_elasticache_clusters
from modules.msk import list_kafka_clusters
from modules.sg import list_security_groups
from modules.sg_detail import fetch_all_sg_data
from modules.route53 import list_route53, fetch_route53_data, sanitize_sheet_name
from modules.opensearch import list_opensearch_clusters
from modules.dynamodb import list_dynamodb_tables
from modules.eks import list_eks_clusters
from modules.acm import list_acm_certificates
from modules.kms import list_kms_keys
from modules.secrets_manager import list_secrets_manager
from modules.sqs import list_sqs_queues
from modules.ses import list_ses_identities
from modules.sns import list_sns_topics
from modules.lamda import list_lambda_functions
from openpyxl.worksheet.table import Table, TableStyleInfo

app = Flask(__name__)

RESOURCE_MAP = {
    "vpcs": list_vpcs,
    "subnets": list_subnets,
    "eks": list_eks_clusters,
    "asg": list_auto_scaling_groups,
    "ec2": list_ec2_instances,
    "security-groups": list_security_groups,
    "nacl": list_nacls,
    "elbs": list_elbs,
    "target-groups": list_target_groups,
    "database": list_db_clusters,
    "dynamodb": list_dynamodb_tables,
    "elasticache": list_elasticache_clusters,
    "msk": list_kafka_clusters,
    "opensearch": list_opensearch_clusters,
    "route53": list_route53,
    "cloudfront": list_cloudfront_distributions,
    "s3": list_s3_buckets,
    "lamda": list_lambda_functions,
    "acm": list_acm_certificates,
    "kms": list_kms_keys,
    "secrets-manager": list_secrets_manager,
    "sqs": list_sqs_queues,
    "ses": list_ses_identities,
    "sns": list_sns_topics
}

DETAIL_RESOURCE_MAP = {
    "security-groups-details": fetch_all_sg_data,
    "route53-details": fetch_route53_data
}

REGION_LIST = [
    "us-east-1",    # N. Virginia
    "us-east-2",    # Ohio
    "us-west-1",    # N. California
    "us-west-2",    # Oregon
    "ap-south-1",   # Mumbai
    "ap-northeast-3", # Osaka
    "ap-northeast-2", # Seoul
    "ap-southeast-1", # Singapore
    "ap-southeast-2", # Sydney
    "ap-northeast-1", # Tokyo
    "ca-central-1", # Canada Central
    "eu-central-1", # Frankfurt
    "eu-west-1",    # Ireland
    "eu-west-2",    # London
    "eu-west-3",    # Paris
    "eu-north-1",   # Stockholm
    "sa-east-1"     # São Paulo
]

def get_aws_profiles(config_path="~/.aws/config"):
    profiles = []
    path = os.path.expanduser(config_path)
    if os.path.exists(path):
        with open(path, 'r') as f:
            content = f.read()
            profiles = re.findall(r'\[profile\s+([^\]]+)\]', content)

    return sorted(profiles)

def has_valid_sso_token(config_path="~/.aws/config", cache_dir="~/.aws/sso/cache"):
    config = configparser.ConfigParser()
    config.read(os.path.expanduser(config_path))

    sso_session = next(
        (section for section in config.sections() if section.startswith("sso-session ")),
        None
    )

    if not sso_session:
        print("[ERROR] No sso-session section found in config.")
        return False

    start_url = config[sso_session].get("sso_start_url")
    region = config[sso_session].get("sso_region")

    if not start_url or not region:
        print(f"[ERROR] sso_start_url or sso_region missing in '{sso_session}'.")
        return False

    cache_path = os.path.expanduser(cache_dir)
    if not os.path.isdir(cache_path):
        print("[INFO] Cache directory not found.")
        return False

    for filename in os.listdir(cache_path):
        try:
            with open(os.path.join(cache_path, filename), "r") as f:
                data = json.load(f)

                # Debugging
                print(f"[DEBUG] Checking cache file {filename}")
                print(f"        startUrl: {data.get('startUrl')}")
                print(f"        region: {data.get('region')}")
                print(f"        expiresAt: {data.get('expiresAt')}")

                # startUrl, region
                if data.get("startUrl") != start_url or data.get("region") != region:
                    continue

                # expiresAt 
                expires_at = data.get("expiresAt")
                if not expires_at:
                    continue

                expires = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
                if expires > datetime.now(timezone.utc):
                    return True  # valid session
        except Exception as e:
            print(f"[WARN] Failed to parse {filename}: {e}")
            continue

    return False  # no valid session

def save_excel_with_format(sheet_dict, filename):
    def sanitize_datetime(df):
        for col in df.select_dtypes(include=['datetimetz', 'datetime64[ns, UTC]']):
            df[col] = df[col].dt.tz_localize(None)
        return df

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in sheet_dict.items():
            df = sanitize_datetime(pd.DataFrame(df))
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(filename)
    header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Column width auto adjust + center alignment
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create a Table
        if ws.max_row > 1 and ws.max_column > 0:
            table_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            safe_table_name = f"{sheet.replace(' ', '_').replace('-', '_')}_Table"
            table = Table(displayName=safe_table_name[:30], ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

    # Hosted Zones → Hyperlink to other sheets
    if "Hosted Zones" in wb.sheetnames:
        summary_ws = wb["Hosted Zones"]
        for row in summary_ws.iter_rows(min_row=2, max_col=1):
            cell = row[0]
            zone_name = str(cell.value).rstrip('.')
            target_sheet = sanitize_sheet_name(zone_name)[:31]
            if target_sheet in wb.sheetnames:
                cell.hyperlink = f"#{quote_sheetname(target_sheet)}!A1"
                cell.style = "Hyperlink"

    # Add each Zone sheet → Hosted Zones hyperlink
    for sheet in wb.sheetnames:
        if sheet == "Hosted Zones":
            continue
        ws = wb[sheet]
        last_row = ws.max_row + 2

        # Merge & Write Text
        max_col = ws.max_column
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=max_col)

        cell = ws.cell(row=last_row, column=1)
        cell.value = "Go to Hosted Zones"
        cell.hyperlink = f"#{quote_sheetname('Hosted Zones')}!A1"
        cell.style = "Hyperlink"

        # Background color & Alignment
        cell.fill = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)

def parallel_execute(resource_map, session):
    results = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(func, session): key for key, func in resource_map.items()}
        for future in as_completed(futures):
            key = futures[future]
            try:
                results[key] = future.result()
            except Exception as e:
                print(f"[ERROR] {key} failed: {e}")
                results[key] = []
    return results

def save_excel_from_data(data_dict, sheet_order):
    
    def sanitize_datetime(df):
        for col in df.select_dtypes(include=['datetime64[ns, UTC]', 'datetimetz']):
            df[col] = df[col].dt.tz_localize(None)
        return df
    
    def workbook_with_format(file_name_with_dir):
        try:
            workbook = load_workbook(file_name_with_dir)
            header_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.auto_filter.ref = worksheet.dimensions
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
                    column_letter = get_column_letter(column[0].column)
                    worksheet.column_dimensions[column_letter].width = max_length
                    for cell in column:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            workbook.save(file_name_with_dir)
        except Exception as e:
            print(f"Error formatting workbook: {e}")

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            sheet_written = False
            for sheet_name in sheet_order:
                if sheet_name not in data_dict:
                    continue
                data = data_dict[sheet_name]

                if isinstance(data, tuple):
                    sub_sheets = ["Summary", "Details", "Findings"]
                    for sub_sheet, df in zip(sub_sheets, data):
                        df = sanitize_datetime(pd.DataFrame(df))
                        if df.empty:
                            print(f"[WARNING] Sheet {sheet_name}-{sub_sheet} has no data, skipping.")
                            continue
                        sheet_written = True
                        df.to_excel(writer, sheet_name=sub_sheet, index=False)
                else:
                    df = sanitize_datetime(pd.DataFrame(data))
                    if df.empty:
                        print(f"[WARNING] Sheet {sheet_name} has no data, skipping.")
                        continue
                    sheet_written = True
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            if not sheet_written:
                print("[WARNING] No sheets written. Excel file will be empty.")

        workbook_with_format(tmp.name)
        return tmp.name

@app.route('/api/<resource>')
def get_resource(resource):
    if resource not in RESOURCE_MAP:
        return jsonify({"error": "Unsupported resource"}), 404

    profile = request.args.get("profile", "sightmind-prod")
    region = request.args.get("region", "us-east-1")
    try:
        session = boto3.Session(profile_name=profile, region_name=region)
        result = RESOURCE_MAP[resource](session)
        print(f"[DEBUG] {resource} result: {result}")
        if not result:
            return jsonify({"columns": [], "rows": []})
        
        columns = list(result[0].keys())
        rows = [list(item.values()) for item in result]
        return jsonify({"columns": columns, "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download/selected')
def selected_region_download():
    profile = request.args.get("profile")
    region = request.args.get("region")
    if not profile or not region:
        return "Profile & Region is required", 400

    try:
        session = boto3.Session(profile_name=profile, region_name=region)

        # Collect Data
        inventory_data = parallel_execute(RESOURCE_MAP, session)
        detail_data = parallel_execute(DETAIL_RESOURCE_MAP, session)

        excel_files = []

        # # Inventory Excel
        if any(inventory_data.values()):
            inventory_excel = save_excel_from_data(inventory_data, sheet_order=list(RESOURCE_MAP.keys()))
            excel_files.append((inventory_excel, f"{profile}_inventory.xlsx"))
        else:
            print("[INFO] No Inventory data found. Skipping Inventory Excel.")

        # SG Detail Excel
        sg_raw = detail_data.get("security-groups-details", [])
        if sg_raw:
            sg_data = {"security-groups-details": sg_raw}
            sg_excel = save_excel_from_data(sg_data, sheet_order=["security-groups-details"])
            excel_files.append((sg_excel, f"{profile}_detail_sg.xlsx"))
        else:
            print("[INFO] No Security Group data found. Skipping SG Excel.")

        # Route53 Excel
        route53_data = detail_data.get("route53-details", {})
        has_route53_data = any(not df.empty for df in route53_data.values()) if route53_data else False

        if has_route53_data:
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as route53_tmp:
                save_excel_with_format(route53_data, route53_tmp.name)
                excel_files.append((route53_tmp.name, f"{profile}_route53.xlsx"))
        else:
            print("[INFO] No Route53 data found in any sheet. Skipping Route53 Excel.")

        # ZIP Compression
        with NamedTemporaryFile(delete=False, suffix=".zip") as tmp_zip:
            with zipfile.ZipFile(tmp_zip.name, 'w') as zipf:
                if not excel_files:
                    print("[WARNING] No Excel files generated. Returning empty zip.")
                for file_path, arc_name in excel_files:
                    zipf.write(file_path, arcname=arc_name)

        # Delete temporary Excel file
        for file_path, _ in excel_files:
            os.remove(file_path)

        return send_file(tmp_zip.name, download_name=f"{profile}_aws_inventory_{datetime.now().strftime('%y_%m_%d')}.zip", as_attachment=True)

    except Exception as e:
        return str(e), 500

@app.route('/')
def index():
    profiles = get_aws_profiles()
    return render_template('index.html', resource_keys=list(RESOURCE_MAP.keys()),
                            profile_list=profiles,
                            profile_len=len(profiles),
                            sso_valid=has_valid_sso_token(),
                            region_list=REGION_LIST,
                        )

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)